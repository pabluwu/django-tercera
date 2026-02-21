import os
from django.core.management.base import BaseCommand, CommandError
from django.contrib.auth import get_user_model
from django.db import transaction
from openpyxl import load_workbook

from bomberos.models import UserProfile


class Command(BaseCommand):
    help = "Importa datos de Registro_LITE.xlsx a User y UserProfile."

    HEADER_MAP = {
        "NOMBRES": "nombres",
        "APELLIDO PATERNO": "apellido_paterno",
        "APELLIDO MATERNO": "apellido_materno",
        "CIA": "cia",
        "RUT": "rut",
        "REGISTRO": "registro",
        "REGISTRO CIA": "registro_cia",
        "CODIGO DE LLAMADO": "codigo_llamado",
        "CARGO": "cargo",
        "E-MAIL": "email",
        "TELEFONO": "telefono",
        "SEXO": "sexo",
        "NACIONALIDAD": "nacionalidad",
        "SANGRE GRUPO": "sangre_grupo",
        "ESTADO CIVIL": "estado_civil",
        "PROFESION": "profesion",
        "DIRECCION CALLE": "direccion_calle",
        "DIRECCION NUMERO": "direccion_numero",
        "DIRECCION COMPLEMENTO": "direccion_complemento",
        "DIRECCION COMUNA": "direccion_comuna",
    }

    def add_arguments(self, parser):
        parser.add_argument(
            "file_path",
            nargs="?",
            default="Registro_LITE.xlsx",
            help="Ruta al archivo Registro_LITE.xlsx (por defecto en el root del proyecto)",
        )
        parser.add_argument(
            "--password-from-rut",
            action="store_true",
            help="Si se indica, establece la contraseña igual al rut (texto literal).",
        )

    def handle(self, *args, **options):
        file_path = options["file_path"]
        set_password_from_rut = options["password_from_rut"]
        if not os.path.exists(file_path):
            raise CommandError(f"No se encontró el archivo: {file_path}")

        wb = load_workbook(file_path, read_only=True)
        sheet = wb.active

        headers = [str(cell.value).strip() if cell.value is not None else "" for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        header_index = {name: idx for idx, name in enumerate(headers)}

        missing_headers = [h for h in self.HEADER_MAP if h not in header_index]
        if missing_headers:
            raise CommandError(f"Faltan columnas requeridas: {', '.join(missing_headers)}")

        User = get_user_model()
        created_users = 0
        updated_users = 0
        created_profiles = 0
        updated_profiles = 0

        for row in sheet.iter_rows(min_row=2, values_only=True):
            # Skip completely empty rows
            if not any(row):
                continue

            values = {}
            for header, model_field in self.HEADER_MAP.items():
                idx = header_index.get(header)
                values[model_field] = row[idx] if idx is not None else None

            rut = (values.get("rut") or "").strip()
            if not rut:
                self.stdout.write(self.style.WARNING("Fila omitida por no tener RUT"))
                continue

            email = (values.get("email") or "").strip()
            nombres = (values.get("nombres") or "").strip()
            ap_pat = (values.get("apellido_paterno") or "").strip()
            ap_mat = (values.get("apellido_materno") or "").strip()
            full_last_name = " ".join([p for p in [ap_pat, ap_mat] if p]).strip()

            with transaction.atomic():
                user, created = User.objects.get_or_create(
                    username=rut,
                    defaults={
                        "email": email,
                        "first_name": nombres,
                        "last_name": full_last_name,
                    },
                )
                if created:
                    created_users += 1
                    if set_password_from_rut:
                        user.set_password(rut)
                    else:
                        user.set_unusable_password()
                else:
                    updated = False
                    if email and user.email != email:
                        user.email = email
                        updated = True
                    if nombres and user.first_name != nombres:
                        user.first_name = nombres
                        updated = True
                    if full_last_name and user.last_name != full_last_name:
                        user.last_name = full_last_name
                        updated = True
                    if updated:
                        updated_users += 1
                user.save()

                profile, p_created = UserProfile.objects.get_or_create(user=user)
                for field, value in values.items():
                    if field == "email":
                        continue  # no es parte del perfil
                    if field == "telefono":
                        try:
                            clean_value = int(value) if value not in (None, "") else None
                        except (TypeError, ValueError):
                            clean_value = None
                        setattr(profile, field, clean_value)
                    else:
                        setattr(profile, field, value or "")
                if p_created:
                    created_profiles += 1
                else:
                    updated_profiles += 1
                profile.save()

        self.stdout.write(
            self.style.SUCCESS(
                f"Importación completada. Usuarios creados: {created_users}, actualizados: {updated_users}. "
                f"Perfiles creados: {created_profiles}, actualizados: {updated_profiles}."
            )
        )
